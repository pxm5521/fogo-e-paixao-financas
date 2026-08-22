#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Recalcula a taxonomia de categorias/subcategorias diretamente das colunas
da aba ExtratoNu — CORRIGIDO: Tipo (=categoria) e Classe (=subcategoria).
(Pedro corrigiu: primeiro eu tinha entendido Classe=categoria/Tipo=subcategoria,
o certo é o inverso.)

Gera:
  - categories_data.json   (para eu escrever src/lib/categories.js)
  - seed_transactions.json (substitui o antigo, agora com categoriaId/
    subcategoriaId = slug(Tipo)/slug(Classe) direto da planilha, sem
    heurística de palavra-chave)
  - seed_report.json
"""
import json
import sys
import hashlib
import datetime
import unicodedata
from pathlib import Path
from collections import defaultdict
import openpyxl

if len(sys.argv) < 2:
    print("Uso: python3 generate_seed.py caminho/para/Saldo_em_Conta_FP.xlsx")
    sys.exit(1)

SRC = sys.argv[1]
HERE = Path(__file__).resolve().parent
TODAY = datetime.date(2026, 8, 22)

# Variantes de grafia (case/acento) que são claramente o mesmo valor —
# mapeadas para a forma mais usada. Aplicadas ao valor bruto de cada coluna
# (independente de qual vira categoria/subcategoria).
CLASSE_MERGE = {
    "ensaio": "Ensaio",
    "som": "Som",
    "pedro": "Pedro",
    "participacao especial": "Participação Especial",
}

TIPO_MERGE = {
    "investimento": "Investimento",
    "carnaval 2025": "Carnaval 2025",
    "carnaval 2026": "Carnaval 2026",
}


def clean(s):
    if s is None:
        return ""
    return " ".join(str(s).strip().split())


def norm_key(s):
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower()


def slugify(s):
    key = norm_key(s).strip()
    key = "".join(c if c.isalnum() else "-" for c in key)
    while "--" in key:
        key = key.replace("--", "-")
    return key.strip("-")


def canonical_tipo(raw):
    """Coluna Tipo -> vira CATEGORIA. Vazio cai em 'Sem Categoria'."""
    cleaned = clean(raw)
    if not cleaned:
        return "Sem Categoria"
    return TIPO_MERGE.get(norm_key(cleaned), cleaned)


def canonical_classe(raw):
    """Coluna Classe -> vira SUBCATEGORIA. Vazio = sem subcategoria."""
    cleaned = clean(raw)
    if not cleaned:
        return None
    return CLASSE_MERGE.get(norm_key(cleaned), cleaned)


def parse_date(d):
    if isinstance(d, datetime.datetime):
        return d.date()
    if isinstance(d, str):
        try:
            return datetime.datetime.strptime(d.strip(), "%d/%m/%Y").date()
        except ValueError:
            return None
    return None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["ExtratoNu"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    # categoriaId -> {label, tipos:set, subcategorias: {subId: {label, count}}, count}
    categories = defaultdict(lambda: {"label": "", "tipos": set(), "subcategorias": {}, "count": 0})

    rows_out = []
    n_sem_categoria = 0
    n_sem_subcategoria = 0
    skipped_no_value = 0
    skipped_future = 0

    for rownum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        valor = row[idx["R$"]]
        if valor in (None, ""):
            skipped_no_value += 1
            continue
        date = parse_date(row[idx["Data"]])
        if date is None:
            continue
        if date > TODAY:
            skipped_future += 1
            continue

        quem = row[idx.get("Quem")]
        motivo = row[idx.get("Motivo")]
        classe_raw = row[idx.get("Classe")]
        tipo_raw = row[idx.get("Tipo")]

        quem = clean(quem) or None
        motivo = clean(motivo) or None

        tipo_lanc = "receita" if valor > 0 else "despesa"

        categoria = canonical_tipo(tipo_raw)       # Tipo -> categoria
        subcategoria = canonical_classe(classe_raw)  # Classe -> subcategoria

        cat_id = slugify(categoria)
        sub_id = slugify(subcategoria) if subcategoria else None

        cat = categories[cat_id]
        cat["label"] = categoria
        cat["tipos"].add(tipo_lanc)
        cat["count"] += 1
        if sub_id:
            sub = cat["subcategorias"].setdefault(sub_id, {"label": subcategoria, "count": 0})
            sub["count"] += 1

        if categoria == "Sem Categoria":
            n_sem_categoria += 1
        elif not sub_id:
            n_sem_subcategoria += 1

        raw_key = f"{date.isoformat()}|{valor}|{quem}|{motivo}|{rownum}"
        uid = "hist_" + hashlib.sha1(raw_key.encode("utf-8")).hexdigest()[:20]
        descricao = " - ".join([p for p in [quem, motivo] if p]) or "(sem descricao)"

        rows_out.append({
            "id": uid,
            "data": date.isoformat(),
            "valor": round(float(valor), 2),
            "tipo": tipo_lanc,
            "descricao": descricao,
            "quem": quem,
            "motivoOriginal": motivo,
            "categoriaId": cat_id,
            "subcategoriaId": sub_id,
            "evento": None,
            "origem": "historico_planilha",
            "classificacaoAutomatica": categoria != "Sem Categoria",
            "revisado": categoria != "Sem Categoria",
        })

    rows_out.sort(key=lambda t: t["data"])

    # Monta a lista final de categorias (ordem alfabética; "Sem Categoria" por último)
    cat_list = []
    for cat_id, cat in categories.items():
        subs = sorted(
            ({"id": sid, "label": s["label"], "count": s["count"]} for sid, s in cat["subcategorias"].items()),
            key=lambda s: s["label"].lower(),
        )
        cat_list.append({
            "id": cat_id,
            "label": cat["label"],
            "tipos": sorted(cat["tipos"]),
            "subcategorias": subs,
            "count": cat["count"],
        })
    cat_list.sort(key=lambda c: (c["label"] == "Sem Categoria", c["label"].lower()))

    with open(HERE / "categories_data.json", "w", encoding="utf-8") as f:
        json.dump(cat_list, f, ensure_ascii=False, indent=2)

    with open(HERE / "seed_transactions.json", "w", encoding="utf-8") as f:
        json.dump(rows_out, f, ensure_ascii=False, indent=2)

    report = {
        "total_lancamentos": len(rows_out),
        "total_categorias": len(cat_list),
        "sem_categoria": n_sem_categoria,
        "com_categoria_sem_subcategoria": n_sem_subcategoria,
        "linhas_ignoradas_sem_valor": skipped_no_value,
        "linhas_ignoradas_data_futura": skipped_future,
        "periodo": [rows_out[0]["data"], rows_out[-1]["data"]] if rows_out else None,
    }
    with open(HERE / "seed_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"\n{len(cat_list)} categorias geradas.")


if __name__ == "__main__":
    main()
